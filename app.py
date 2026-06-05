import os
from dotenv import load_dotenv
load_dotenv()

import re
import math
import uuid
from io import BytesIO
from datetime import datetime, timedelta, timezone

AR_TZ = timezone(timedelta(hours=-3))

def ar_now():
    return datetime.now(AR_TZ)

def to_ar(dt):
    """Convert a naive UTC datetime (from DB) to Argentina time."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(AR_TZ)
from functools import wraps

import psycopg2
import psycopg2.extras
from flask import Flask, render_template, request, redirect, url_for, jsonify, session, abort, Response, stream_with_context
from minio import Minio
from minio.error import S3Error
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'valuaciones-sarmiento-2024-xk9')
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024  # 20 MB

DATABASE_URL  = os.environ.get('DATABASE_URL', '')
MINIO_ENDPOINT   = os.environ.get('MINIO_ENDPOINT',   '62.72.11.137:9000')
MINIO_ACCESS_KEY = os.environ.get('MINIO_ACCESS_KEY', 'minioadmin')
MINIO_SECRET_KEY = os.environ.get('MINIO_SECRET_KEY', 'kzkvmhwlhrjbebbt')
MINIO_BUCKET     = os.environ.get('MINIO_BUCKET',     'valuaciones')
MINIO_SECURE     = os.environ.get('MINIO_SECURE', 'false').lower() == 'true'

minio_client = Minio(
    MINIO_ENDPOINT,
    access_key=MINIO_ACCESS_KEY,
    secret_key=MINIO_SECRET_KEY,
    secure=MINIO_SECURE,
)

USUARIOS_INICIALES = [
    ('Mariano', 'Sarmiento302', 'usuario'),
    ('Luis',    'Sarmiento302', 'usuario'),
    ('Octavio', 'Sarmiento302', 'usuario'),
]

ALLOWED_EXTENSIONS = {'.pdf', '.doc', '.docx', '.xls', '.xlsx', '.jpg', '.jpeg', '.png'}
PER_PAGE = 20


# ---- DB helpers ----

def get_db():
    return psycopg2.connect(DATABASE_URL)


def fetchall(conn, sql, params=None):
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(sql, params or [])
        return cur.fetchall()


def fetchone(conn, sql, params=None):
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute(sql, params or [])
        return cur.fetchone()


def fetchscalar(conn, sql, params=None):
    with conn.cursor() as cur:
        cur.execute(sql, params or [])
        row = cur.fetchone()
        return row[0] if row else None


def execute(conn, sql, params=None):
    with conn.cursor() as cur:
        cur.execute(sql, params or [])


def obtener_siguiente_vr(conn, usuario):
    return fetchscalar(conn, '''
        INSERT INTO contadores_vr (usuario, ultimo_vr)
        VALUES (%s, 1)
        ON CONFLICT (usuario) DO UPDATE
        SET ultimo_vr = contadores_vr.ultimo_vr + 1
        RETURNING ultimo_vr
    ''', (usuario,))


def obtener_siguiente_recuento(conn, usuario):
    return fetchscalar(conn, '''
        INSERT INTO contadores_recuento (usuario, ultimo_recuento)
        VALUES (%s, 1)
        ON CONFLICT (usuario) DO UPDATE
        SET ultimo_recuento = contadores_recuento.ultimo_recuento + 1
        RETURNING ultimo_recuento
    ''', (usuario,))


def parse_pesos(val):
    if not val:
        return 0
    try:
        cleaned = str(val).replace('.', '').replace(',', '').replace('$', '').replace(' ', '')
        return int(cleaned)
    except (ValueError, TypeError):
        return 0


# ---- MinIO ----

def init_minio():
    try:
        if not minio_client.bucket_exists(MINIO_BUCKET):
            minio_client.make_bucket(MINIO_BUCKET)
    except S3Error as e:
        print(f'[MinIO] Error al inicializar bucket: {e}')


# ---- DB init ----

def init_db():
    conn = get_db()

    # ---- Tabla expedientes ----
    execute(conn, '''
        CREATE TABLE IF NOT EXISTS expedientes (
            id                SERIAL PRIMARY KEY,
            expediente        TEXT,
            caratula          TEXT,
            creado_por        TEXT,
            activa            INTEGER DEFAULT 1,
            eliminado_por     TEXT,
            fecha_eliminacion TEXT,
            fecha_registro    TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # ---- Tabla contadores_vr ----
    execute(conn, '''
        CREATE TABLE IF NOT EXISTS contadores_vr (
            usuario TEXT PRIMARY KEY,
            ultimo_vr INTEGER DEFAULT 0
        )
    ''')

    # ---- Tabla catastros (id interno, numero_vr visible) ----
    execute(conn, '''
        CREATE TABLE IF NOT EXISTS catastros (
            id                          SERIAL PRIMARY KEY,
            expediente_id               INTEGER NOT NULL REFERENCES expedientes(id),
            numero_vr                   INTEGER,
            catastro                    TEXT,
            tipo_catastro               TEXT,
            direccion                   TEXT,
            fecha                       TEXT,
            terreno_m2                  REAL DEFAULT 0,
            terreno_frente_lado         TEXT,
            terreno_antes_revision      TEXT,
            usd_m2_terreno              REAL DEFAULT 0,
            productiva_hect             REAL DEFAULT 0,
            productiva_frente_lado      TEXT,
            productiva_antes_revision   TEXT,
            usd_hect_productiva         REAL DEFAULT 0,
            con_monte_hect              REAL DEFAULT 0,
            con_monte_frente_lado       TEXT,
            con_monte_antes_revision    TEXT,
            usd_hect_con_monte          REAL DEFAULT 0,
            cerros_hect                 REAL DEFAULT 0,
            cerros_frente_lado          TEXT,
            cerros_antes_revision       TEXT,
            usd_hect_cerros             REAL DEFAULT 0,
            sup_edif_m2                 REAL DEFAULT 0,
            edif_frente_lado            TEXT,
            edif_antes_revision         TEXT,
            usd_m2_edif                 REAL DEFAULT 0,
            valor_dolar                 REAL DEFAULT 0,
            total_usd_terreno           REAL DEFAULT 0,
            total_usd_edif              REAL DEFAULT 0,
            total_usd                   REAL DEFAULT 0,
            propuesta                   REAL DEFAULT 0,
            monto                       REAL DEFAULT 0,
            denuncia                    TEXT,
            gmaps_zona                  TEXT,
            gmaps_frente                TEXT,
            terreno_total               REAL DEFAULT 0,
            fot                         REAL DEFAULT 0,
            fos                         REAL DEFAULT 0,
            sup_edif_total              REAL DEFAULT 0,
            pisos_maximos               INTEGER DEFAULT 0,
            porcentaje_emprendimiento   REAL DEFAULT 0,
            costo_usd_m2_emprendimiento REAL DEFAULT 0,
            emprendimiento              REAL DEFAULT 0,
            observaciones               TEXT,
            latitud                     REAL,
            longitud                    REAL,
            editado_por                 TEXT,
            fecha_registro              TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            es_reconsideracion          BOOLEAN DEFAULT FALSE
        )
    ''')

    # Columnas nuevas en catastros para bases ya existentes
    for col, definition in [
        ('productiva_frente_lado',    'TEXT'),
        ('productiva_antes_revision', 'TEXT'),
        ('con_monte_frente_lado',     'TEXT'),
        ('con_monte_antes_revision',  'TEXT'),
        ('cerros_frente_lado',        'TEXT'),
        ('cerros_antes_revision',     'TEXT'),
        ('es_reconsideracion',        'BOOLEAN DEFAULT FALSE'),
        ('numero_vr',                 'INTEGER'),
    ]:
        execute(conn, f'ALTER TABLE catastros ADD COLUMN IF NOT EXISTS {col} {definition}')



    # ---- Tabla archivos ----
    execute(conn, '''
        CREATE TABLE IF NOT EXISTS archivos (
            id               SERIAL PRIMARY KEY,
            catastro_id      INTEGER NOT NULL,
            nombre_original  TEXT NOT NULL,
            objeto_minio     TEXT NOT NULL,
            tipo             TEXT,
            tamanio          BIGINT DEFAULT 0,
            subido_por       TEXT,
            fecha_subida     TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Si la tabla archivos ya existía con valuacion_id, agregar catastro_id y migrar
    execute(conn, 'ALTER TABLE archivos ADD COLUMN IF NOT EXISTS catastro_id INTEGER')
    old_col = fetchscalar(conn,
        "SELECT COUNT(*) FROM information_schema.columns "
        "WHERE table_name='archivos' AND column_name='valuacion_id'")
    if old_col:
        execute(conn, 'UPDATE archivos SET catastro_id = valuacion_id WHERE catastro_id IS NULL')
        execute(conn, 'ALTER TABLE archivos DROP COLUMN IF EXISTS valuacion_id')

    # ---- Tabla automotores ----
    execute(conn, '''
        CREATE TABLE IF NOT EXISTS automotores (
            id               SERIAL PRIMARY KEY,
            expediente_id    INTEGER NOT NULL REFERENCES expedientes(id),
            numero_recuento  INTEGER,
            vehiculo         TEXT,
            anio             INTEGER,
            valor            BIGINT DEFAULT 0,
            fecha            TEXT,
            creado_por       TEXT,
            observaciones    TEXT,
            fecha_registro   TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # ---- Tabla contadores_recuento ----
    execute(conn, '''
        CREATE TABLE IF NOT EXISTS contadores_recuento (
            usuario TEXT PRIMARY KEY,
            ultimo_recuento INTEGER DEFAULT 0
        )
    ''')

    # automotor_id en archivos
    execute(conn, 'ALTER TABLE archivos ADD COLUMN IF NOT EXISTS automotor_id INTEGER')

    # Columnas de reconsideración en automotores
    execute(conn, 'ALTER TABLE automotores ADD COLUMN IF NOT EXISTS es_reconsideracion BOOLEAN DEFAULT FALSE')
    execute(conn, 'ALTER TABLE automotores ADD COLUMN IF NOT EXISTS editado_por TEXT')

    # Columnas de eliminación lógica
    execute(conn, 'ALTER TABLE catastros ADD COLUMN IF NOT EXISTS eliminado BOOLEAN DEFAULT FALSE')
    execute(conn, 'ALTER TABLE automotores ADD COLUMN IF NOT EXISTS eliminado BOOLEAN DEFAULT FALSE')

    # Moneda de propuesta y monto en catastros
    execute(conn, "ALTER TABLE catastros ADD COLUMN IF NOT EXISTS propuesta_moneda TEXT DEFAULT 'USD'")
    execute(conn, "ALTER TABLE catastros ADD COLUMN IF NOT EXISTS monto_moneda TEXT DEFAULT 'ARS'")

    # Cotización dólar y moneda del valor en automotores
    execute(conn, 'ALTER TABLE automotores ADD COLUMN IF NOT EXISTS cotizacion_dolar NUMERIC DEFAULT 0')
    execute(conn, "ALTER TABLE automotores ADD COLUMN IF NOT EXISTS valor_moneda TEXT DEFAULT 'ARS'")

    # Otro link adicional en catastros
    execute(conn, 'ALTER TABLE catastros ADD COLUMN IF NOT EXISTS otro_link TEXT')

    # Departamento en catastros
    execute(conn, 'ALTER TABLE catastros ADD COLUMN IF NOT EXISTS departamento TEXT')

    # ---- Tabla usuarios ----
    execute(conn, '''
        CREATE TABLE IF NOT EXISTS usuarios (
            id            SERIAL PRIMARY KEY,
            username      TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL
        )
    ''')
    execute(conn, "ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS rol TEXT DEFAULT 'usuario'")

    # Admin: siempre sincronizar contraseña y rol
    admin_hash = generate_password_hash('admin123')
    if fetchone(conn, 'SELECT id FROM usuarios WHERE username = %s', ('admin',)):
        execute(conn, "UPDATE usuarios SET password_hash = %s, rol = 'admin' WHERE username = %s",
                (admin_hash, 'admin'))
    else:
        execute(conn, "INSERT INTO usuarios (username, password_hash, rol) VALUES ('admin', %s, 'admin')",
                (admin_hash,))

    for username, password, rol in USUARIOS_INICIALES:
        if not fetchone(conn, 'SELECT id FROM usuarios WHERE username = %s', (username,)):
            execute(conn, 'INSERT INTO usuarios (username, password_hash, rol) VALUES (%s, %s, %s)',
                    (username, generate_password_hash(password), rol))
        else:
            execute(conn, "UPDATE usuarios SET rol = %s WHERE username = %s AND (rol IS NULL OR rol = '')",
                    (rol, username))

    conn.commit()
    conn.close()
    init_minio()


# ---- Auth ----

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


@app.template_filter('filesize')
def filesize_filter(size):
    if not size:
        return '—'
    if size < 1024:
        return f'{size} B'
    elif size < 1024 * 1024:
        return f'{size / 1024:.1f} KB'
    return f'{size / 1024 / 1024:.1f} MB'


@app.template_filter('usd')
def usd_filter(value):
    if not value:
        return '0'
    return f'{float(value):,.0f}'.replace(',', '.')


@app.template_filter('pesos')
def pesos_filter(value):
    if not value:
        return '0'
    return f'{int(value):,}'.replace(',', '.')


# ---- Helpers ----

def haversine(lat1, lon1, lat2, lon2):
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return 2 * 6371 * math.asin(math.sqrt(a))


def extraer_coordenadas(url):
    if not url:
        return None, None
    lat_matches = re.findall(r'!3d(-?[\d.]+)', url)
    lon_matches = re.findall(r'!4d(-?[\d.]+)', url)
    if lat_matches and lon_matches:
        return float(lat_matches[-1]), float(lon_matches[-1])
    for pattern in [r'q=(-?\d+\.?\d*),(-?\d+\.?\d*)', r'll=(-?\d+\.?\d*),(-?\d+\.?\d*)', r'@(-?\d+\.?\d*),(-?\d+\.?\d*)']:
        m = re.search(pattern, url)
        if m:
            return float(m.group(1)), float(m.group(2))
    return None, None


def parse_float(val):
    try:
        return float(val) if val else 0
    except (ValueError, TypeError):
        return 0


def parse_int(val):
    try:
        return int(val) if val else 0
    except (ValueError, TypeError):
        return 0


def _calcular_catastro(data):
    """Calcula los totales a partir del form data. Devuelve dict con los valores calculados."""
    tipo_catastro  = data.get('tipo_catastro', '').strip()
    sup_edif_m2    = parse_float(data.get('sup_edif_m2'))
    usd_m2_edif    = parse_float(data.get('usd_m2_edif'))
    terreno_total  = parse_float(data.get('terreno_total'))
    fot            = parse_float(data.get('fot'))
    fos            = parse_float(data.get('fos'))
    pisos_maximos  = parse_int(data.get('pisos_maximos'))
    valor_dolar    = parse_float(data.get('valor_dolar'))
    sup_edif_total = parse_float(data.get('sup_edif_total_calc'))
    porcentaje_emprendimiento   = parse_float(data.get('porcentaje_emprendimiento'))
    costo_usd_m2_emprendimiento = parse_float(data.get('costo_usd_m2_emprendimiento'))
    emprendimiento = parse_float(data.get('emprendimiento'))

    if tipo_catastro == 'Rural':
        terreno_m2     = 0
        usd_m2_terreno = 0
        productiva_hect             = parse_float(data.get('productiva_hect'))
        productiva_frente_lado      = data.get('productiva_frente_lado', '').strip()
        productiva_antes_revision   = data.get('productiva_antes_revision', '').strip()
        usd_hect_productiva         = parse_float(data.get('usd_hect_productiva'))
        con_monte_hect              = parse_float(data.get('con_monte_hect'))
        con_monte_frente_lado       = data.get('con_monte_frente_lado', '').strip()
        con_monte_antes_revision    = data.get('con_monte_antes_revision', '').strip()
        usd_hect_con_monte          = parse_float(data.get('usd_hect_con_monte'))
        cerros_hect                 = parse_float(data.get('cerros_hect'))
        cerros_frente_lado          = data.get('cerros_frente_lado', '').strip()
        cerros_antes_revision       = data.get('cerros_antes_revision', '').strip()
        usd_hect_cerros             = parse_float(data.get('usd_hect_cerros'))
        total_usd_terreno = (productiva_hect * usd_hect_productiva +
                             con_monte_hect   * usd_hect_con_monte  +
                             cerros_hect      * usd_hect_cerros)
    else:
        terreno_m2     = parse_float(data.get('terreno_m2'))
        usd_m2_terreno = parse_float(data.get('usd_m2_terreno'))
        productiva_hect = usd_hect_productiva = 0
        productiva_frente_lado = productiva_antes_revision = ''
        con_monte_hect  = usd_hect_con_monte  = 0
        con_monte_frente_lado = con_monte_antes_revision = ''
        cerros_hect     = usd_hect_cerros     = 0
        cerros_frente_lado = cerros_antes_revision = ''
        total_usd_terreno = terreno_m2 * usd_m2_terreno

    total_usd_edif = sup_edif_m2 * usd_m2_edif
    total_usd      = total_usd_terreno + total_usd_edif
    propuesta      = parse_float(data.get('propuesta')) or total_usd * 1.10
    propuesta_moneda = data.get('propuesta_moneda', 'USD') or 'USD'
    monto_moneda     = data.get('monto_moneda', 'ARS') or 'ARS'

    gmaps_zona   = data.get('gmaps_zona', '').strip()
    gmaps_frente = data.get('gmaps_frente', '').strip()
    lat, lon = extraer_coordenadas(gmaps_frente or gmaps_zona)

    dep_sel   = data.get('departamento_sel', '').strip()
    dep_otro  = data.get('departamento_otro', '').strip()
    departamento = dep_otro if dep_sel == 'Otros' else dep_sel

    return dict(
        tipo_catastro=tipo_catastro,
        catastro=data.get('catastro', '').strip(),
        direccion=data.get('direccion', '').strip(),
        fecha=data.get('fecha', ''),
        terreno_m2=terreno_m2,
        terreno_frente_lado=data.get('terreno_frente_lado', '').strip(),
        terreno_antes_revision=data.get('terreno_antes_revision', '').strip(),
        usd_m2_terreno=usd_m2_terreno,
        productiva_hect=productiva_hect,
        productiva_frente_lado=productiva_frente_lado,
        productiva_antes_revision=productiva_antes_revision,
        usd_hect_productiva=usd_hect_productiva,
        con_monte_hect=con_monte_hect,
        con_monte_frente_lado=con_monte_frente_lado,
        con_monte_antes_revision=con_monte_antes_revision,
        usd_hect_con_monte=usd_hect_con_monte,
        cerros_hect=cerros_hect,
        cerros_frente_lado=cerros_frente_lado,
        cerros_antes_revision=cerros_antes_revision,
        usd_hect_cerros=usd_hect_cerros,
        sup_edif_m2=sup_edif_m2,
        edif_frente_lado=data.get('edif_frente_lado', '').strip(),
        edif_antes_revision=data.get('edif_antes_revision', '').strip(),
        usd_m2_edif=usd_m2_edif,
        valor_dolar=valor_dolar,
        total_usd_terreno=total_usd_terreno, total_usd_edif=total_usd_edif,
        total_usd=total_usd, propuesta=propuesta,
        propuesta_moneda=propuesta_moneda,
        monto=parse_float(data.get('monto')),
        monto_moneda=monto_moneda,
        denuncia=data.get('denuncia', '').strip(),
        gmaps_zona=gmaps_zona, gmaps_frente=gmaps_frente,
        otro_link=data.get('otro_link', '').strip(),
        departamento=departamento,
        terreno_total=terreno_total, fot=fot, fos=fos,
        sup_edif_total=sup_edif_total, pisos_maximos=pisos_maximos,
        porcentaje_emprendimiento=porcentaje_emprendimiento,
        costo_usd_m2_emprendimiento=costo_usd_m2_emprendimiento,
        emprendimiento=emprendimiento,
        observaciones=data.get('observaciones', '').strip(),
        latitud=lat, longitud=lon,
    )


# ---- Routes ----

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'usuario' in session:
        return redirect(url_for('index'))
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        user = fetchone(conn, 'SELECT * FROM usuarios WHERE username = %s', (username,))
        conn.close()
        if user and check_password_hash(user['password_hash'], password):
            session['usuario'] = username
            session['rol']     = user.get('rol') or 'usuario'
            return redirect(url_for('index'))
        error = 'Usuario o contraseña incorrectos.'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.pop('usuario', None)
    session.pop('rol', None)
    return redirect(url_for('login'))


@app.context_processor
def inject_globals():
    return {'is_admin': session.get('rol') == 'admin'}


@app.route('/')
@login_required
def index():
    q               = request.args.get('q', '').strip()
    desde           = request.args.get('desde', '').strip()
    hasta           = request.args.get('hasta', '').strip()
    filtro_usuario  = request.args.get('usuario', '').strip()
    ver_eliminados  = request.args.get('ver_eliminados', '') == '1' and session.get('rol') == 'admin'
    page            = max(1, int(request.args.get('page', 1) or 1))

    conditions = [] if ver_eliminados else ['e.activa = 1']
    params = []
    if q:
        conditions.append('''(
            c.catastro ILIKE %s OR e.expediente ILIKE %s OR e.caratula ILIKE %s
            OR c.direccion ILIKE %s OR c.denuncia ILIKE %s
        )''')
        params.extend([f'%{q}%'] * 5)
    if desde:
        conditions.append('c.fecha >= %s')
        params.append(desde)
    if hasta:
        conditions.append('c.fecha <= %s')
        params.append(hasta)
    if filtro_usuario:
        conditions.append('e.creado_por = %s')
        params.append(filtro_usuario)

    where = ' AND '.join(conditions)

    if q or desde or hasta:
        join_type = 'INNER JOIN'
        conditions.append('(c.eliminado IS NOT TRUE)')
        where = ' AND '.join(conditions)
    else:
        join_type = 'LEFT JOIN'

    conn = get_db()

    count_sql = f'''
        SELECT COUNT(DISTINCT e.id)
        FROM expedientes e
        {join_type} catastros c ON c.expediente_id = e.id
        WHERE {where}
    '''
    total = fetchscalar(conn, count_sql, params)

    list_sql = f'''
        SELECT e.id, e.expediente, e.caratula, e.creado_por,
               e.fecha_registro, e.activa,
               COUNT(CASE WHEN c.es_reconsideracion = FALSE AND (c.eliminado IS NOT TRUE) THEN 1 END) AS num_catastros,
               (SELECT COUNT(*) FROM automotores WHERE expediente_id = e.id AND (eliminado IS NOT TRUE)) AS num_automotores,
               MIN(c.fecha) AS primera_fecha,
               MAX(c.fecha) AS ultima_fecha,
               STRING_AGG(CAST(c.numero_vr AS TEXT), ', ' ORDER BY c.id) AS vr_numeros
        FROM expedientes e
        {join_type} catastros c ON c.expediente_id = e.id
        WHERE {where}
        GROUP BY e.id, e.expediente, e.caratula, e.creado_por, e.fecha_registro, e.activa
        ORDER BY e.id DESC
        LIMIT %s OFFSET %s
    '''
    expedientes = fetchall(conn, list_sql, params + [PER_PAGE, (page - 1) * PER_PAGE])
    usuarios_db = fetchall(conn, 'SELECT username FROM usuarios ORDER BY username')
    conn.close()

    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    return render_template('index.html',
                           expedientes=expedientes,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           usuario=session['usuario'],
                           q=q, desde=desde, hasta=hasta,
                           filtro_usuario=filtro_usuario,
                           ver_eliminados=ver_eliminados,
                           page=page, total_pages=total_pages, total=total,
                           usuarios_lista=[u['username'] for u in usuarios_db])


@app.route('/expediente/<int:exp_id>')
@login_required
def expediente_detail(exp_id):
    conn = get_db()
    exp = fetchone(conn, 'SELECT * FROM expedientes WHERE id = %s', (exp_id,))
    if not exp:
        conn.close()
        return redirect(url_for('index'))
    es_admin = session.get('rol') == 'admin'
    filtro_eliminado = '' if es_admin else 'AND (eliminado IS NOT TRUE)'
    catastros = fetchall(conn,
        f'SELECT * FROM catastros WHERE expediente_id = %s {filtro_eliminado} ORDER BY id ASC', (exp_id,))
    automotores_list = fetchall(conn,
        f'SELECT * FROM automotores WHERE expediente_id = %s {filtro_eliminado} ORDER BY id ASC', (exp_id,))
    conn.close()
    return render_template('expediente.html',
                           exp=exp,
                           catastros=catastros,
                           automotores=automotores_list,
                           usuario=session['usuario'])


@app.route('/verificar_catastro', methods=['POST'])
@login_required
def verificar_catastro():
    data = request.get_json()
    catastro    = data.get('catastro', '').strip()
    gmaps_zona  = data.get('gmaps_zona', '').strip()
    gmaps_frente = data.get('gmaps_frente', '').strip()
    exclude_id  = data.get('exclude_id')

    alertas = []
    conn = get_db()

    if catastro:
        sql = '''
            SELECT c.id, e.expediente, e.caratula, c.direccion
            FROM catastros c
            JOIN expedientes e ON e.id = c.expediente_id
            WHERE c.catastro = %s AND e.activa = 1
        '''
        params = [catastro]
        if exclude_id:
            sql += ' AND c.id != %s'
            params.append(exclude_id)
        existente = fetchone(conn, sql, params)
        if existente:
            alertas.append({
                'tipo': 'duplicado',
                'mensaje': f'El catastro {catastro} ya fue tasado (VR #{existente["id"]}) - '
                           f'Expediente: {existente["expediente"]}, '
                           f'Carátula: {existente["caratula"]}, '
                           f'Dirección: {existente["direccion"]}'
            })

    gmaps_link = gmaps_frente or gmaps_zona
    lat, lon = extraer_coordenadas(gmaps_link)

    if lat is not None and lon is not None:
        sql = '''
            SELECT c.id, c.catastro, e.expediente, e.caratula, c.direccion, c.latitud, c.longitud
            FROM catastros c
            JOIN expedientes e ON e.id = c.expediente_id
            WHERE c.latitud IS NOT NULL AND c.longitud IS NOT NULL AND e.activa = 1
        '''
        params = []
        if exclude_id:
            sql += ' AND c.id != %s'
            params.append(exclude_id)
        registros = fetchall(conn, sql, params)
        for reg in registros:
            dist = haversine(lat, lon, reg['latitud'], reg['longitud'])
            if dist < 0.5:
                dist_metros = round(dist * 1000)
                alertas.append({
                    'tipo': 'proximidad',
                    'mensaje': f'El catastro {reg["catastro"]} (VR #{reg["id"]}) está a {dist_metros}m - '
                               f'Expediente: {reg["expediente"]}, '
                               f'Carátula: {reg["caratula"]}, '
                               f'Dirección: {reg["direccion"]}',
                    'distancia': dist_metros
                })

    conn.close()
    return jsonify({'alertas': alertas, 'coordenadas': {'lat': lat, 'lon': lon}})


@app.route('/reconsideraciones')
@login_required
def reconsideraciones():
    desde          = request.args.get('desde', '').strip()
    hasta          = request.args.get('hasta', '').strip()
    filtro_usuario = request.args.get('usuario', '').strip()

    conditions = ['c.es_reconsideracion = TRUE', '(c.eliminado IS NOT TRUE)']
    params = []
    if desde:
        conditions.append('c.fecha >= %s')
        params.append(desde)
    if hasta:
        conditions.append('c.fecha <= %s')
        params.append(hasta)
    if filtro_usuario:
        conditions.append('c.editado_por = %s')
        params.append(filtro_usuario)

    where = ' AND '.join(conditions)

    conn = get_db()
    rows = fetchall(conn, f'''
        SELECT c.id, c.numero_vr, c.catastro, c.tipo_catastro,
               c.direccion, c.fecha, c.editado_por,
               e.id AS exp_id, e.expediente, e.caratula
        FROM catastros c
        JOIN expedientes e ON e.id = c.expediente_id
        WHERE {where}
        ORDER BY c.fecha DESC, c.id DESC
    ''', params)
    usuarios_db = fetchall(conn, 'SELECT username FROM usuarios ORDER BY username')
    conn.close()

    return render_template('reconsideraciones.html',
                           rows=rows,
                           desde=desde,
                           hasta=hasta,
                           filtro_usuario=filtro_usuario,
                           usuarios_lista=[u['username'] for u in usuarios_db],
                           usuario=session['usuario'])


@app.route('/nuevo_expediente')
@login_required
def nuevo_expediente():
    return render_template('nuevo_expediente.html', usuario=session['usuario'])


@app.route('/crear_expediente', methods=['POST'])
@login_required
def crear_expediente():
    usuario_actual = session['usuario']
    expediente_num = request.form.get('expediente', '').strip()
    caratula = request.form.get('caratula', '').strip()

    conn = get_db()
    exp_id = fetchscalar(conn, '''
        INSERT INTO expedientes (expediente, caratula, creado_por, activa)
        VALUES (%s, %s, %s, 1) RETURNING id
    ''', (expediente_num, caratula, usuario_actual))
    conn.commit()
    conn.close()
    return redirect(url_for('expediente_detail', exp_id=exp_id))


@app.route('/expediente/<int:exp_id>/nuevo_catastro')
@login_required
def nuevo_catastro(exp_id):
    conn = get_db()
    exp = fetchone(conn, 'SELECT * FROM expedientes WHERE id = %s', (exp_id,))
    conn.close()
    if not exp:
        return redirect(url_for('index'))
    return render_template('form.html',
                           expediente_obj=exp,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           usuario=session['usuario'])


@app.route('/expediente/<int:exp_id>/guardar_catastro', methods=['POST'])
@login_required
def guardar_catastro(exp_id):
    """Agrega un catastro a un expediente existente."""
    usuario_actual = session['usuario']
    data = request.form

    conn = get_db()
    exp = fetchone(conn, 'SELECT id FROM expedientes WHERE id = %s', (exp_id,))
    if not exp:
        conn.close()
        return redirect(url_for('index'))

    c = _calcular_catastro(data)
    num_vr = obtener_siguiente_vr(conn, usuario_actual)
    cat_id = fetchscalar(conn, '''
        INSERT INTO catastros (
            expediente_id,
            numero_vr,
            catastro, tipo_catastro, direccion, fecha,
            terreno_m2, terreno_frente_lado, terreno_antes_revision, usd_m2_terreno,
            productiva_hect, productiva_frente_lado, productiva_antes_revision, usd_hect_productiva,
            con_monte_hect,  con_monte_frente_lado,  con_monte_antes_revision,  usd_hect_con_monte,
            cerros_hect,     cerros_frente_lado,     cerros_antes_revision,     usd_hect_cerros,
            sup_edif_m2, edif_frente_lado, edif_antes_revision, usd_m2_edif,
            valor_dolar,
            total_usd_terreno, total_usd_edif, total_usd, propuesta, propuesta_moneda, monto, monto_moneda,
            denuncia, gmaps_zona, gmaps_frente, otro_link, departamento,
            terreno_total, fot, fos, sup_edif_total, pisos_maximos,
            porcentaje_emprendimiento, costo_usd_m2_emprendimiento, emprendimiento,
            observaciones, latitud, longitud
        ) VALUES (
            %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s,
            %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s
        ) RETURNING id
    ''', (
        exp_id, num_vr,
        c['catastro'], c['tipo_catastro'], c['direccion'], c['fecha'],
        c['terreno_m2'], c['terreno_frente_lado'], c['terreno_antes_revision'], c['usd_m2_terreno'],
        c['productiva_hect'], c['productiva_frente_lado'], c['productiva_antes_revision'], c['usd_hect_productiva'],
        c['con_monte_hect'],  c['con_monte_frente_lado'],  c['con_monte_antes_revision'],  c['usd_hect_con_monte'],
        c['cerros_hect'],     c['cerros_frente_lado'],     c['cerros_antes_revision'],     c['usd_hect_cerros'],
        c['sup_edif_m2'], c['edif_frente_lado'], c['edif_antes_revision'], c['usd_m2_edif'],
        c['valor_dolar'],
        c['total_usd_terreno'], c['total_usd_edif'], c['total_usd'], c['propuesta'], c['propuesta_moneda'], c['monto'], c['monto_moneda'],
        c['denuncia'], c['gmaps_zona'], c['gmaps_frente'], c['otro_link'], c['departamento'],
        c['terreno_total'], c['fot'], c['fos'], c['sup_edif_total'], c['pisos_maximos'],
        c['porcentaje_emprendimiento'], c['costo_usd_m2_emprendimiento'], c['emprendimiento'],
        c['observaciones'], c['latitud'], c['longitud'],
    ))

    for file in request.files.getlist('archivos'):
        if not file or not file.filename:
            continue
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            continue
        objeto_minio = f'{cat_id}/{uuid.uuid4().hex}{ext}'
        file_data = file.read()
        tamanio = len(file_data)
        try:
            minio_client.put_object(MINIO_BUCKET, objeto_minio, BytesIO(file_data), tamanio,
                                    content_type=file.content_type or 'application/octet-stream')
            execute(conn, '''
                INSERT INTO archivos (catastro_id, nombre_original, objeto_minio, tipo, tamanio, subido_por)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (cat_id, file.filename, objeto_minio, ext, tamanio, usuario_actual))
        except S3Error as e:
            print(f'[MinIO] Error al subir archivo: {e}')

    conn.commit()
    conn.close()
    return redirect(url_for('expediente_detail', exp_id=exp_id))


@app.route('/desactivar/<int:exp_id>', methods=['POST'])
@login_required
def desactivar(exp_id):
    conn = get_db()
    execute(conn,
        'UPDATE expedientes SET activa = 0, eliminado_por = %s, fecha_eliminacion = %s WHERE id = %s',
        (session['usuario'], datetime.now().strftime('%Y-%m-%d %H:%M'), exp_id))
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for('index'))


@app.route('/reactivar/<int:exp_id>', methods=['POST'])
@login_required
def reactivar(exp_id):
    conn = get_db()
    execute(conn,
        'UPDATE expedientes SET activa = 1, eliminado_por = NULL, fecha_eliminacion = NULL WHERE id = %s',
        (exp_id,))
    conn.commit()
    conn.close()
    return redirect(request.referrer or url_for('index'))


@app.route('/eliminar_catastro/<int:cat_id>', methods=['POST'])
@login_required
def eliminar_catastro(cat_id):
    conn = get_db()
    cat = fetchone(conn, 'SELECT expediente_id FROM catastros WHERE id = %s', (cat_id,))
    if not cat:
        conn.close()
        return redirect(url_for('index'))
    execute(conn, 'UPDATE catastros SET eliminado = TRUE WHERE id = %s', (cat_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('expediente_detail', exp_id=cat['expediente_id']))


@app.route('/ver/<int:cat_id>')
@login_required
def ver(cat_id):
    conn = get_db()
    catastro = fetchone(conn, 'SELECT * FROM catastros WHERE id = %s', (cat_id,))
    if not catastro:
        conn.close()
        return redirect(url_for('index'))
    exp = fetchone(conn, 'SELECT * FROM expedientes WHERE id = %s', (catastro['expediente_id'],))
    archivos = fetchall(conn, 'SELECT * FROM archivos WHERE catastro_id = %s ORDER BY fecha_subida DESC', (cat_id,))
    conn.close()
    return render_template('form.html',
                           viendo=catastro,
                           expediente_obj=exp,
                           archivos=archivos,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           usuario=session['usuario'])


@app.route('/editar_catastro/<int:cat_id>')
@login_required
def editar_catastro(cat_id):
    conn = get_db()
    catastro = fetchone(conn, 'SELECT * FROM catastros WHERE id = %s', (cat_id,))
    if not catastro:
        conn.close()
        return redirect(url_for('index'))
    exp = fetchone(conn, 'SELECT * FROM expedientes WHERE id = %s', (catastro['expediente_id'],))
    conn.close()
    return render_template('form.html',
                           editando=catastro,
                           expediente_obj=exp,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           usuario=session['usuario'])


@app.route('/guardar_editar_catastro/<int:cat_id>', methods=['POST'])
@login_required
def guardar_editar_catastro(cat_id):
    usuario_actual = session['usuario']
    data = request.form

    conn = get_db()
    cat = fetchone(conn, 'SELECT expediente_id FROM catastros WHERE id = %s', (cat_id,))
    if not cat:
        conn.close()
        return redirect(url_for('index'))
    exp_id = cat['expediente_id']

    c = _calcular_catastro(data)
    execute(conn, '''
        UPDATE catastros SET
            catastro=%s, tipo_catastro=%s, direccion=%s, fecha=%s,
            terreno_m2=%s, terreno_frente_lado=%s, terreno_antes_revision=%s, usd_m2_terreno=%s,
            productiva_hect=%s, productiva_frente_lado=%s, productiva_antes_revision=%s, usd_hect_productiva=%s,
            con_monte_hect=%s, con_monte_frente_lado=%s, con_monte_antes_revision=%s, usd_hect_con_monte=%s,
            cerros_hect=%s, cerros_frente_lado=%s, cerros_antes_revision=%s, usd_hect_cerros=%s,
            sup_edif_m2=%s, edif_frente_lado=%s, edif_antes_revision=%s, usd_m2_edif=%s,
            valor_dolar=%s,
            total_usd_terreno=%s, total_usd_edif=%s, total_usd=%s,
            propuesta=%s, propuesta_moneda=%s, monto=%s, monto_moneda=%s,
            denuncia=%s, gmaps_zona=%s, gmaps_frente=%s, otro_link=%s, departamento=%s,
            terreno_total=%s, fot=%s, fos=%s, sup_edif_total=%s, pisos_maximos=%s,
            porcentaje_emprendimiento=%s, costo_usd_m2_emprendimiento=%s, emprendimiento=%s,
            observaciones=%s, latitud=%s, longitud=%s, editado_por=%s
        WHERE id=%s
    ''', (
        c['catastro'], c['tipo_catastro'], c['direccion'], c['fecha'],
        c['terreno_m2'], c['terreno_frente_lado'], c['terreno_antes_revision'], c['usd_m2_terreno'],
        c['productiva_hect'], c['productiva_frente_lado'], c['productiva_antes_revision'], c['usd_hect_productiva'],
        c['con_monte_hect'], c['con_monte_frente_lado'], c['con_monte_antes_revision'], c['usd_hect_con_monte'],
        c['cerros_hect'], c['cerros_frente_lado'], c['cerros_antes_revision'], c['usd_hect_cerros'],
        c['sup_edif_m2'], c['edif_frente_lado'], c['edif_antes_revision'], c['usd_m2_edif'],
        c['valor_dolar'],
        c['total_usd_terreno'], c['total_usd_edif'], c['total_usd'],
        c['propuesta'], c['propuesta_moneda'], c['monto'], c['monto_moneda'],
        c['denuncia'], c['gmaps_zona'], c['gmaps_frente'], c['otro_link'], c['departamento'],
        c['terreno_total'], c['fot'], c['fos'], c['sup_edif_total'], c['pisos_maximos'],
        c['porcentaje_emprendimiento'], c['costo_usd_m2_emprendimiento'], c['emprendimiento'],
        c['observaciones'], c['latitud'], c['longitud'], usuario_actual,
        cat_id,
    ))
    conn.commit()
    conn.close()
    return redirect(url_for('expediente_detail', exp_id=exp_id))


@app.route('/reconsiderar/<int:cat_id>')
@login_required
def reconsiderar(cat_id):
    conn = get_db()
    catastro = fetchone(conn, 'SELECT * FROM catastros WHERE id = %s', (cat_id,))
    if not catastro:
        conn.close()
        return redirect(url_for('index'))
    exp = fetchone(conn, 'SELECT * FROM expedientes WHERE id = %s', (catastro['expediente_id'],))
    archivos = fetchall(conn, 'SELECT * FROM archivos WHERE catastro_id = %s ORDER BY fecha_subida DESC', (cat_id,))
    conn.close()
    
    # Mover valores viejos a "Antes / Revisión"
    catastro['terreno_antes_revision'] = str(catastro.get('usd_m2_terreno', '')) if catastro.get('usd_m2_terreno') else ''
    catastro['productiva_antes_revision'] = str(catastro.get('usd_hect_productiva', '')) if catastro.get('usd_hect_productiva') else ''
    catastro['con_monte_antes_revision'] = str(catastro.get('usd_hect_con_monte', '')) if catastro.get('usd_hect_con_monte') else ''
    catastro['cerros_antes_revision'] = str(catastro.get('usd_hect_cerros', '')) if catastro.get('usd_hect_cerros') else ''
    catastro['edif_antes_revision'] = str(catastro.get('usd_m2_edif', '')) if catastro.get('usd_m2_edif') else ''

    return render_template('form.html',
                           reconsiderando=catastro,
                           expediente_obj=exp,
                           archivos=archivos,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           usuario=session['usuario'])


@app.route('/guardar_reconsideracion/<int:cat_id>', methods=['POST'])
@login_required
def guardar_reconsideracion(cat_id):
    usuario_actual = session['usuario']
    data = request.form

    conn = get_db()
    catastro_viejo = fetchone(conn, 'SELECT expediente_id FROM catastros WHERE id = %s', (cat_id,))
    if not catastro_viejo:
        conn.close()
        return redirect(url_for('index'))
    exp_id = catastro_viejo['expediente_id']

    # Actualizar expediente si se enviaron esos campos
    new_expediente = data.get('expediente', '').strip()
    new_caratula   = data.get('caratula', '').strip()
    if new_expediente or new_caratula:
        execute(conn, '''
            UPDATE expedientes SET
                expediente = COALESCE(NULLIF(%s,''), expediente),
                caratula   = COALESCE(NULLIF(%s,''), caratula)
            WHERE id = %s
        ''', (new_expediente, new_caratula, exp_id))

    c = _calcular_catastro(data)
    num_vr = obtener_siguiente_vr(conn, usuario_actual)
    nuevo_cat_id = fetchscalar(conn, '''
        INSERT INTO catastros (
            expediente_id,
            numero_vr,
            catastro, tipo_catastro, direccion, fecha,
            terreno_m2, terreno_frente_lado, terreno_antes_revision, usd_m2_terreno,
            productiva_hect, productiva_frente_lado, productiva_antes_revision, usd_hect_productiva,
            con_monte_hect,  con_monte_frente_lado,  con_monte_antes_revision,  usd_hect_con_monte,
            cerros_hect,     cerros_frente_lado,     cerros_antes_revision,     usd_hect_cerros,
            sup_edif_m2, edif_frente_lado, edif_antes_revision, usd_m2_edif,
            valor_dolar,
            total_usd_terreno, total_usd_edif, total_usd, propuesta, propuesta_moneda, monto, monto_moneda,
            denuncia, gmaps_zona, gmaps_frente, otro_link, departamento,
            terreno_total, fot, fos, sup_edif_total, pisos_maximos,
            porcentaje_emprendimiento, costo_usd_m2_emprendimiento, emprendimiento,
            observaciones, latitud, longitud, editado_por, es_reconsideracion
        ) VALUES (
            %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s,
            %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s,
            %s, %s, %s, %s, TRUE
        ) RETURNING id
    ''', (
        exp_id, num_vr,
        c['catastro'], c['tipo_catastro'], c['direccion'], c['fecha'],
        c['terreno_m2'], c['terreno_frente_lado'], c['terreno_antes_revision'], c['usd_m2_terreno'],
        c['productiva_hect'], c['productiva_frente_lado'], c['productiva_antes_revision'], c['usd_hect_productiva'],
        c['con_monte_hect'],  c['con_monte_frente_lado'],  c['con_monte_antes_revision'],  c['usd_hect_con_monte'],
        c['cerros_hect'],     c['cerros_frente_lado'],     c['cerros_antes_revision'],     c['usd_hect_cerros'],
        c['sup_edif_m2'], c['edif_frente_lado'], c['edif_antes_revision'], c['usd_m2_edif'],
        c['valor_dolar'],
        c['total_usd_terreno'], c['total_usd_edif'], c['total_usd'], c['propuesta'], c['propuesta_moneda'], c['monto'], c['monto_moneda'],
        c['denuncia'], c['gmaps_zona'], c['gmaps_frente'], c['otro_link'], c['departamento'],
        c['terreno_total'], c['fot'], c['fos'], c['sup_edif_total'], c['pisos_maximos'],
        c['porcentaje_emprendimiento'], c['costo_usd_m2_emprendimiento'], c['emprendimiento'],
        c['observaciones'], c['latitud'], c['longitud'], usuario_actual
    ))

    # Copiar archivos adjuntos del catastro viejo al nuevo
    archivos_viejos = fetchall(conn, 'SELECT * FROM archivos WHERE catastro_id = %s', (cat_id,))
    for arch in archivos_viejos:
        execute(conn, '''
            INSERT INTO archivos (catastro_id, nombre_original, objeto_minio, tipo, tamanio, subido_por)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (nuevo_cat_id, arch['nombre_original'], arch['objeto_minio'], arch['tipo'], arch['tamanio'], arch['subido_por']))

    for file in request.files.getlist('archivos'):
        if not file or not file.filename:
            continue
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            continue
        objeto_minio = f'{nuevo_cat_id}/{uuid.uuid4().hex}{ext}'
        file_data = file.read()
        tamanio = len(file_data)
        try:
            minio_client.put_object(MINIO_BUCKET, objeto_minio, BytesIO(file_data), tamanio,
                                    content_type=file.content_type or 'application/octet-stream')
            execute(conn, '''
                INSERT INTO archivos (catastro_id, nombre_original, objeto_minio, tipo, tamanio, subido_por)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (nuevo_cat_id, file.filename, objeto_minio, ext, tamanio, usuario_actual))
        except S3Error as e:
            print(f'[MinIO] Error al subir archivo: {e}')

    conn.commit()
    conn.close()
    return redirect(url_for('expediente_detail', exp_id=exp_id))


# ---- Archivos (MinIO) ----

@app.route('/subir_archivo/<int:catastro_id>', methods=['POST'])
@login_required
def subir_archivo(catastro_id):
    conn = get_db()
    cat = fetchone(conn, 'SELECT expediente_id FROM catastros WHERE id = %s', (catastro_id,))
    if not cat:
        conn.close()
        return redirect(url_for('index'))

    if 'archivo' not in request.files:
        conn.close()
        return redirect(url_for('editar', cat_id=catastro_id))

    file = request.files['archivo']
    if not file.filename:
        conn.close()
        return redirect(url_for('editar', cat_id=catastro_id))

    nombre_original = file.filename
    ext = os.path.splitext(nombre_original)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        conn.close()
        return redirect(url_for('editar', cat_id=catastro_id))

    objeto_minio = f'{catastro_id}/{uuid.uuid4().hex}{ext}'
    data = file.read()
    tamanio = len(data)

    try:
        minio_client.put_object(
            MINIO_BUCKET, objeto_minio, BytesIO(data), tamanio,
            content_type=file.content_type or 'application/octet-stream',
        )
    except S3Error as e:
        print(f'[MinIO] Error al subir archivo: {e}')
        conn.close()
        return redirect(url_for('editar', cat_id=catastro_id))

    execute(conn, '''
        INSERT INTO archivos (catastro_id, nombre_original, objeto_minio, tipo, tamanio, subido_por)
        VALUES (%s, %s, %s, %s, %s, %s)
    ''', (catastro_id, nombre_original, objeto_minio, ext, tamanio, session['usuario']))
    conn.commit()
    conn.close()
    return redirect(url_for('editar', cat_id=catastro_id))


@app.route('/archivo/<int:id>')
@login_required
def descargar_archivo(id):
    conn = get_db()
    archivo = fetchone(conn, 'SELECT * FROM archivos WHERE id = %s', (id,))
    conn.close()
    if not archivo:
        abort(404)
    try:
        obj = minio_client.get_object(MINIO_BUCKET, archivo['objeto_minio'])
        ext = (archivo['tipo'] or '').lower()
        disposition = 'inline' if ext in ('.jpg', '.jpeg', '.png', '.pdf') else f'attachment; filename="{archivo["nombre_original"]}"'
        content_type = obj.headers.get('content-type', 'application/octet-stream')
        return Response(
            stream_with_context(obj.stream(32 * 1024)),
            content_type=content_type,
            headers={'Content-Disposition': disposition},
        )
    except S3Error as e:
        print(f'[MinIO] Error al servir archivo: {e}')
        abort(500)


@app.route('/eliminar_archivo/<int:id>', methods=['POST'])
@login_required
def eliminar_archivo(id):
    conn = get_db()
    archivo = fetchone(conn, 'SELECT * FROM archivos WHERE id = %s', (id,))
    if not archivo:
        conn.close()
        return redirect(url_for('index'))
    try:
        minio_client.remove_object(MINIO_BUCKET, archivo['objeto_minio'])
    except S3Error as e:
        print(f'[MinIO] Error al eliminar objeto: {e}')
    execute(conn, 'DELETE FROM archivos WHERE id = %s', (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('editar', cat_id=archivo['catastro_id']))


@app.route('/exportar_excel')
@login_required
def exportar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    q              = request.args.get('q', '').strip()
    desde          = request.args.get('desde', '').strip()
    hasta          = request.args.get('hasta', '').strip()
    filtro_usuario = request.args.get('usuario', '').strip()

    conditions = ['e.activa = 1', '(c.eliminado IS NOT TRUE)']
    params = []
    if q:
        conditions.append('(c.catastro ILIKE %s OR e.expediente ILIKE %s OR e.caratula ILIKE %s OR c.direccion ILIKE %s OR c.denuncia ILIKE %s)')
        params.extend([f'%{q}%'] * 5)
    if desde:
        conditions.append('c.fecha >= %s')
        params.append(desde)
    if hasta:
        conditions.append('c.fecha <= %s')
        params.append(hasta)
    if filtro_usuario:
        conditions.append('e.creado_por = %s')
        params.append(filtro_usuario)

    conn = get_db()
    rows = fetchall(conn, f'''
        SELECT c.id, e.expediente, c.catastro, c.tipo_catastro, e.caratula,
               c.direccion, c.fecha, c.denuncia,
               c.terreno_m2, c.sup_edif_m2, c.usd_m2_terreno, c.usd_m2_edif,
               c.valor_dolar, c.total_usd_terreno, c.total_usd_edif, c.total_usd,
               c.propuesta, c.fot, c.fos, c.sup_edif_total, c.pisos_maximos,
               c.emprendimiento, e.creado_por, c.fecha_registro, c.observaciones
        FROM catastros c
        JOIN expedientes e ON e.id = c.expediente_id
        WHERE {" AND ".join(conditions)}
        ORDER BY c.id DESC
    ''', params)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Valuaciones'

    header_fill = PatternFill('solid', fgColor='1E3A8A')
    header_font = Font(bold=True, color='DBEAFE', size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alt_fill = PatternFill('solid', fgColor='F8F9FB')
    border_side = Side(style='thin', color='D8DCE3')
    cell_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    center = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    usd_fmt = '#,##0.00'

    COLUMNAS = [
        ('VR #',            'id',                           10,  'center'),
        ('Expediente',      'expediente',                   16,  'left'),
        ('Catastro',        'catastro',                     18,  'left'),
        ('Tipo Catastro',   'tipo_catastro',                14,  'center'),
        ('Carátula',        'caratula',                     28,  'left'),
        ('Dirección',       'direccion',                    32,  'left'),
        ('Fecha',           'fecha',                        13,  'center'),
        ('Tipo',            'denuncia',                     13,  'center'),
        ('Terreno m²',      'terreno_m2',                   13,  'right'),
        ('Sup. Edif. m²',   'sup_edif_m2',                  13,  'right'),
        ('U$S/m² Terreno',  'usd_m2_terreno',               15,  'right'),
        ('U$S/m² Edif.',    'usd_m2_edif',                  13,  'right'),
        ('Valor Dólar',     'valor_dolar',                  13,  'right'),
        ('Total U$S Terr.', 'total_usd_terreno',            15,  'right'),
        ('Total U$S Edif.', 'total_usd_edif',               15,  'right'),
        ('Total U$S',       'total_usd',                    14,  'right'),
        ('Propuesta U$S',   'propuesta',                    15,  'right'),
        ('FOT',             'fot',                          8,   'center'),
        ('FOS',             'fos',                          8,   'center'),
        ('Sup. Edif. Total','sup_edif_total',               15,  'right'),
        ('Pisos Máx.',      'pisos_maximos',                10,  'center'),
        ('Emprendimiento',  'emprendimiento',               16,  'right'),
        ('Creado por',      'creado_por',                   14,  'center'),
        ('Fecha Registro',  'fecha_registro',               18,  'center'),
        ('Observaciones',   'observaciones',                40,  'left'),
    ]

    USD_COLS = {'usd_m2_terreno','usd_m2_edif','valor_dolar','total_usd_terreno',
                'total_usd_edif','total_usd','propuesta','emprendimiento','terreno_m2','sup_edif_m2','sup_edif_total'}

    ws.row_dimensions[1].height = 36
    for col_idx, (titulo, _, ancho, _align) in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = cell_border
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    for row_idx, v in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        for col_idx, (_, campo, _, align) in enumerate(COLUMNAS, 1):
            valor = v[campo] if campo in v.keys() else None
            if isinstance(valor, datetime):
                valor = valor.strftime('%d/%m/%Y %H:%M')
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill   = fill
            cell.border = cell_border
            cell.font   = Font(size=9)
            if align == 'center':
                cell.alignment = center
            elif align == 'right':
                cell.alignment = right_align
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=(campo == 'observaciones'))
            if campo in USD_COLS and valor is not None:
                cell.number_format = usd_fmt

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLUMNAS))}1'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fecha_hoy = datetime.now().strftime('%Y%m%d')
    return Response(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="valuaciones_{fecha_hoy}.xlsx"'},
    )


@app.route('/expediente/<int:exp_id>/nuevo_automotor')
@login_required
def nuevo_automotor(exp_id):
    conn = get_db()
    exp = fetchone(conn, 'SELECT * FROM expedientes WHERE id = %s', (exp_id,))
    conn.close()
    if not exp:
        return redirect(url_for('index'))
    return render_template('form_automotor.html',
                           expediente_obj=exp,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           usuario=session['usuario'])


@app.route('/expediente/<int:exp_id>/guardar_automotor', methods=['POST'])
@login_required
def guardar_automotor(exp_id):
    usuario_actual = session['usuario']
    data = request.form

    conn = get_db()
    exp = fetchone(conn, 'SELECT id FROM expedientes WHERE id = %s', (exp_id,))
    if not exp:
        conn.close()
        return redirect(url_for('index'))

    num_recuento = obtener_siguiente_recuento(conn, usuario_actual)
    auto_id = fetchscalar(conn, '''
        INSERT INTO automotores (expediente_id, numero_recuento, vehiculo, anio, valor, fecha, creado_por, observaciones, cotizacion_dolar, valor_moneda)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
    ''', (
        exp_id, num_recuento,
        data.get('vehiculo', '').strip().upper(),
        parse_int(data.get('anio')),
        parse_pesos(data.get('valor')),
        data.get('fecha', ''),
        usuario_actual,
        data.get('observaciones', '').strip(),
        parse_float(data.get('cotizacion_dolar')),
        data.get('valor_moneda', 'ARS') or 'ARS',
    ))

    for file in request.files.getlist('archivos'):
        if not file or not file.filename:
            continue
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            continue
        objeto_minio = f'auto_{auto_id}/{uuid.uuid4().hex}{ext}'
        file_data = file.read()
        tamanio = len(file_data)
        try:
            minio_client.put_object(MINIO_BUCKET, objeto_minio, BytesIO(file_data), tamanio,
                                    content_type=file.content_type or 'application/octet-stream')
            execute(conn, '''
                INSERT INTO archivos (automotor_id, nombre_original, objeto_minio, tipo, tamanio, subido_por)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (auto_id, file.filename, objeto_minio, ext, tamanio, usuario_actual))
        except S3Error as e:
            print(f'[MinIO] Error al subir archivo: {e}')

    conn.commit()
    conn.close()
    return redirect(url_for('expediente_detail', exp_id=exp_id))


@app.route('/automotor/<int:auto_id>')
@login_required
def ver_automotor(auto_id):
    conn = get_db()
    automotor = fetchone(conn, 'SELECT * FROM automotores WHERE id = %s', (auto_id,))
    if not automotor:
        conn.close()
        return redirect(url_for('index'))
    exp = fetchone(conn, 'SELECT * FROM expedientes WHERE id = %s', (automotor['expediente_id'],))
    archivos = fetchall(conn, 'SELECT * FROM archivos WHERE automotor_id = %s ORDER BY fecha_subida DESC', (auto_id,))
    conn.close()
    return render_template('form_automotor.html',
                           viendo=automotor,
                           expediente_obj=exp,
                           archivos=archivos,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           usuario=session['usuario'])


@app.route('/editar_automotor/<int:auto_id>')
@login_required
def editar_automotor(auto_id):
    conn = get_db()
    automotor = fetchone(conn, 'SELECT * FROM automotores WHERE id = %s', (auto_id,))
    if not automotor:
        conn.close()
        return redirect(url_for('index'))
    exp = fetchone(conn, 'SELECT * FROM expedientes WHERE id = %s', (automotor['expediente_id'],))
    conn.close()
    return render_template('form_automotor.html',
                           editando=automotor,
                           expediente_obj=exp,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           usuario=session['usuario'])


@app.route('/guardar_editar_automotor/<int:auto_id>', methods=['POST'])
@login_required
def guardar_editar_automotor(auto_id):
    usuario_actual = session['usuario']
    data = request.form

    conn = get_db()
    auto = fetchone(conn, 'SELECT expediente_id FROM automotores WHERE id = %s', (auto_id,))
    if not auto:
        conn.close()
        return redirect(url_for('index'))
    exp_id = auto['expediente_id']

    execute(conn, '''
        UPDATE automotores SET
            vehiculo=%s, anio=%s, valor=%s, fecha=%s,
            observaciones=%s, cotizacion_dolar=%s, valor_moneda=%s, editado_por=%s
        WHERE id=%s
    ''', (
        data.get('vehiculo', '').strip().upper(),
        parse_int(data.get('anio')),
        parse_pesos(data.get('valor')),
        data.get('fecha', ''),
        data.get('observaciones', '').strip(),
        parse_float(data.get('cotizacion_dolar')),
        data.get('valor_moneda', 'ARS') or 'ARS',
        usuario_actual,
        auto_id,
    ))
    conn.commit()
    conn.close()
    return redirect(url_for('expediente_detail', exp_id=exp_id))


@app.route('/reconsiderar_automotor/<int:auto_id>')
@login_required
def reconsiderar_automotor(auto_id):
    conn = get_db()
    automotor = fetchone(conn, 'SELECT * FROM automotores WHERE id = %s', (auto_id,))
    if not automotor:
        conn.close()
        return redirect(url_for('index'))
    exp = fetchone(conn, 'SELECT * FROM expedientes WHERE id = %s', (automotor['expediente_id'],))
    archivos = fetchall(conn, 'SELECT * FROM archivos WHERE automotor_id = %s ORDER BY fecha_subida DESC', (auto_id,))
    conn.close()
    return render_template('form_automotor.html',
                           reconsiderando=automotor,
                           expediente_obj=exp,
                           archivos=archivos,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           usuario=session['usuario'])


@app.route('/guardar_reconsideracion_automotor/<int:auto_id>', methods=['POST'])
@login_required
def guardar_reconsideracion_automotor(auto_id):
    usuario_actual = session['usuario']
    data = request.form

    conn = get_db()
    auto_viejo = fetchone(conn, 'SELECT expediente_id FROM automotores WHERE id = %s', (auto_id,))
    if not auto_viejo:
        conn.close()
        return redirect(url_for('index'))
    exp_id = auto_viejo['expediente_id']

    num_recuento = obtener_siguiente_recuento(conn, usuario_actual)
    nuevo_auto_id = fetchscalar(conn, '''
        INSERT INTO automotores
            (expediente_id, numero_recuento, vehiculo, anio, valor, fecha,
             creado_por, observaciones, editado_por, es_reconsideracion, cotizacion_dolar, valor_moneda)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE, %s, %s) RETURNING id
    ''', (
        exp_id, num_recuento,
        data.get('vehiculo', '').strip().upper(),
        parse_int(data.get('anio')),
        parse_pesos(data.get('valor')),
        data.get('fecha', ''),
        usuario_actual,
        data.get('observaciones', '').strip(),
        usuario_actual,
        parse_float(data.get('cotizacion_dolar')),
        data.get('valor_moneda', 'ARS') or 'ARS',
    ))

    # Copiar archivos del automotor original
    for arch in fetchall(conn, 'SELECT * FROM archivos WHERE automotor_id = %s', (auto_id,)):
        execute(conn, '''
            INSERT INTO archivos (automotor_id, nombre_original, objeto_minio, tipo, tamanio, subido_por)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (nuevo_auto_id, arch['nombre_original'], arch['objeto_minio'],
              arch['tipo'], arch['tamanio'], arch['subido_por']))

    # Archivos nuevos subidos en la reconsideración
    for file in request.files.getlist('archivos'):
        if not file or not file.filename:
            continue
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in ALLOWED_EXTENSIONS:
            continue
        objeto_minio = f'auto_{nuevo_auto_id}/{uuid.uuid4().hex}{ext}'
        file_data = file.read()
        tamanio = len(file_data)
        try:
            minio_client.put_object(MINIO_BUCKET, objeto_minio, BytesIO(file_data), tamanio,
                                    content_type=file.content_type or 'application/octet-stream')
            execute(conn, '''
                INSERT INTO archivos (automotor_id, nombre_original, objeto_minio, tipo, tamanio, subido_por)
                VALUES (%s, %s, %s, %s, %s, %s)
            ''', (nuevo_auto_id, file.filename, objeto_minio, ext, tamanio, usuario_actual))
        except S3Error as e:
            print(f'[MinIO] Error al subir archivo: {e}')

    conn.commit()
    conn.close()
    return redirect(url_for('expediente_detail', exp_id=exp_id))


@app.route('/automotor/<int:auto_id>/eliminar', methods=['POST'])
@login_required
def eliminar_automotor(auto_id):
    conn = get_db()
    automotor = fetchone(conn, 'SELECT expediente_id FROM automotores WHERE id = %s', (auto_id,))
    if not automotor:
        conn.close()
        return redirect(url_for('index'))
    exp_id = automotor['expediente_id']
    execute(conn, 'UPDATE automotores SET eliminado = TRUE WHERE id = %s', (auto_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('expediente_detail', exp_id=exp_id))


@app.route('/subir_archivo_automotor/<int:auto_id>', methods=['POST'])
@login_required
def subir_archivo_automotor(auto_id):
    conn = get_db()
    auto = fetchone(conn, 'SELECT id FROM automotores WHERE id = %s', (auto_id,))
    if not auto:
        conn.close()
        return redirect(url_for('index'))

    file = request.files.get('archivo')
    if not file or not file.filename:
        conn.close()
        return redirect(url_for('ver_automotor', auto_id=auto_id))

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        conn.close()
        return redirect(url_for('ver_automotor', auto_id=auto_id))

    objeto_minio = f'auto_{auto_id}/{uuid.uuid4().hex}{ext}'
    data = file.read()
    tamanio = len(data)
    try:
        minio_client.put_object(MINIO_BUCKET, objeto_minio, BytesIO(data), tamanio,
                                content_type=file.content_type or 'application/octet-stream')
    except S3Error as e:
        print(f'[MinIO] Error al subir archivo: {e}')
        conn.close()
        return redirect(url_for('ver_automotor', auto_id=auto_id))

    execute(conn, '''
        INSERT INTO archivos (automotor_id, nombre_original, objeto_minio, tipo, tamanio, subido_por)
        VALUES (%s, %s, %s, %s, %s, %s)
    ''', (auto_id, file.filename, objeto_minio, ext, tamanio, session['usuario']))
    conn.commit()
    conn.close()
    return redirect(url_for('ver_automotor', auto_id=auto_id))


@app.route('/eliminar_archivo_automotor/<int:id>', methods=['POST'])
@login_required
def eliminar_archivo_automotor(id):
    conn = get_db()
    archivo = fetchone(conn, 'SELECT * FROM archivos WHERE id = %s', (id,))
    if not archivo:
        conn.close()
        return redirect(url_for('index'))
    try:
        minio_client.remove_object(MINIO_BUCKET, archivo['objeto_minio'])
    except S3Error as e:
        print(f'[MinIO] Error al eliminar objeto: {e}')
    execute(conn, 'DELETE FROM archivos WHERE id = %s', (id,))
    conn.commit()
    conn.close()
    return redirect(url_for('ver_automotor', auto_id=archivo['automotor_id']))


@app.route('/configurar_vr', methods=['GET', 'POST'])
@login_required
def configurar_vr():
    conn = get_db()
    usuario_actual = session['usuario']

    if request.method == 'POST':
        tipo = request.form.get('tipo', 'vr')
        try:
            proximo = int(request.form.get('proximo', 1))
            if proximo < 1:
                proximo = 1
        except Exception:
            proximo = 1

        if tipo == 'recuento':
            execute(conn, '''
                INSERT INTO contadores_recuento (usuario, ultimo_recuento)
                VALUES (%s, %s)
                ON CONFLICT (usuario) DO UPDATE
                SET ultimo_recuento = EXCLUDED.ultimo_recuento
            ''', (usuario_actual, proximo - 1))
        else:
            execute(conn, '''
                INSERT INTO contadores_vr (usuario, ultimo_vr)
                VALUES (%s, %s)
                ON CONFLICT (usuario) DO UPDATE
                SET ultimo_vr = EXCLUDED.ultimo_vr
            ''', (usuario_actual, proximo - 1))
        conn.commit()
        conn.close()
        return redirect(url_for('index'))

    ultimo_vr = fetchscalar(conn, 'SELECT ultimo_vr FROM contadores_vr WHERE usuario = %s', (usuario_actual,))
    ultimo_recuento = fetchscalar(conn, 'SELECT ultimo_recuento FROM contadores_recuento WHERE usuario = %s', (usuario_actual,))
    conn.close()

    proximo_actual = 1 if ultimo_vr is None else ultimo_vr + 1
    proximo_recuento = 1 if ultimo_recuento is None else ultimo_recuento + 1

    return render_template('configurar_vr.html',
                           proximo_actual=proximo_actual,
                           proximo_recuento=proximo_recuento,
                           usuario=usuario_actual)


def _query_reporte(conn, desde, hasta, exp_q, usuario):
    cat_cond, cat_p = [], []
    aut_cond, aut_p = [], []

    if desde:
        cat_cond.append("(c.fecha_registro AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires')::date >= %s"); cat_p.append(desde)
        aut_cond.append("(a.fecha_registro AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires')::date >= %s"); aut_p.append(desde)
    if hasta:
        cat_cond.append("(c.fecha_registro AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires')::date <= %s"); cat_p.append(hasta)
        aut_cond.append("(a.fecha_registro AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires')::date <= %s"); aut_p.append(hasta)
    if exp_q:
        cat_cond.append('e.expediente ILIKE %s'); cat_p.append(f'%{exp_q}%')
        aut_cond.append('e.expediente ILIKE %s'); aut_p.append(f'%{exp_q}%')
    if usuario:
        cat_cond.append('e.creado_por = %s'); cat_p.append(usuario)
        aut_cond.append('a.creado_por = %s'); aut_p.append(usuario)

    cat_cond += ['e.activa = 1', '(c.eliminado IS NOT TRUE)']
    aut_cond += ['e.activa = 1', '(a.eliminado IS NOT TRUE)']
    cat_where = ' AND '.join(cat_cond)
    aut_where = ' AND '.join(aut_cond)

    catastros = fetchall(conn, f'''
        SELECT c.id, c.numero_vr, c.catastro, c.tipo_catastro, c.direccion,
               c.total_usd, c.propuesta, c.monto, c.fecha, c.fecha_registro,
               c.es_reconsideracion, c.editado_por,
               e.id AS exp_id, e.expediente, e.caratula, e.creado_por AS exp_creado_por
        FROM catastros c
        JOIN expedientes e ON e.id = c.expediente_id
        WHERE {cat_where}
        ORDER BY e.id DESC, c.id ASC
    ''', cat_p)

    automotores = fetchall(conn, f'''
        SELECT a.id, a.numero_recuento, a.vehiculo, a.anio, a.valor,
               a.fecha, a.fecha_registro, a.creado_por, a.observaciones,
               a.expediente_id,
               e.expediente, e.caratula, e.creado_por AS exp_creado_por
        FROM automotores a
        JOIN expedientes e ON e.id = a.expediente_id
        WHERE {aut_where}
        ORDER BY a.expediente_id DESC, a.id ASC
    ''', aut_p)

    exp_dict = {}
    for c in catastros:
        eid = c['exp_id']
        if eid not in exp_dict:
            exp_dict[eid] = {'id': eid, 'expediente': c['expediente'],
                             'caratula': c['caratula'], 'creado_por': c['exp_creado_por'],
                             'catastros': [], 'automotores': []}
        exp_dict[eid]['catastros'].append(dict(c))

    for a in automotores:
        eid = a['expediente_id']
        if eid not in exp_dict:
            exp_dict[eid] = {'id': eid, 'expediente': a['expediente'],
                             'caratula': a['caratula'], 'creado_por': a['exp_creado_por'],
                             'catastros': [], 'automotores': []}
        exp_dict[eid]['automotores'].append(dict(a))

    grupos, total_usd, total_pesos = [], 0, 0
    for eid in sorted(exp_dict.keys(), reverse=True):
        g = exp_dict[eid]
        # Para el subtotal tomamos solo el registro más reciente por catastro/vehículo
        # (la reconsideración reemplaza al original, no se suma)
        efectivos_cat = {}
        for c in g['catastros']:
            key = c['catastro'] or str(c['id'])
            if key not in efectivos_cat or c['id'] > efectivos_cat[key]['id']:
                efectivos_cat[key] = c
        sub_usd = sum((c['propuesta'] or 0) for c in efectivos_cat.values())

        efectivos_aut = {}
        for a in g['automotores']:
            key = (a['vehiculo'] or '') + '|' + str(a['anio'] or '')
            if key not in efectivos_aut or a['id'] > efectivos_aut[key]['id']:
                efectivos_aut[key] = a
        sub_pesos = sum((a['valor'] or 0) for a in efectivos_aut.values())
        g['subtotal_usd']   = sub_usd
        g['subtotal_pesos'] = sub_pesos
        total_usd   += sub_usd
        total_pesos += sub_pesos
        grupos.append(g)

    return grupos, total_usd, total_pesos


@app.route('/reporte')
@login_required
def reporte():
    desde   = request.args.get('desde',   '').strip()
    hasta   = request.args.get('hasta',   '').strip()
    exp_q   = request.args.get('exp_q',   '').strip()
    usuario = request.args.get('usuario', '').strip()
    filtrado = bool(desde or hasta or exp_q or usuario)

    conn = get_db()
    grupos, total_usd, total_pesos = ([], 0, 0)
    if filtrado:
        grupos, total_usd, total_pesos = _query_reporte(conn, desde, hasta, exp_q, usuario)
    usuarios_lista = [u['username'] for u in fetchall(conn, 'SELECT username FROM usuarios ORDER BY username')]
    conn.close()

    return render_template('reporte.html',
                           grupos=grupos, total_usd=total_usd, total_pesos=total_pesos,
                           desde=desde, hasta=hasta, exp_q=exp_q, usuario=usuario,
                           filtrado=filtrado, usuarios_lista=usuarios_lista,
                           usuario_actual=session['usuario'])


@app.route('/reporte/excel')
@login_required
def reporte_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    desde   = request.args.get('desde',   '').strip()
    hasta   = request.args.get('hasta',   '').strip()
    exp_q   = request.args.get('exp_q',   '').strip()
    usuario = request.args.get('usuario', '').strip()

    conn = get_db()
    grupos, total_usd, total_pesos = _query_reporte(conn, desde, hasta, exp_q, usuario)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte Consolidado'

    fill_exp    = PatternFill('solid', fgColor='0C4566')
    font_exp    = Font(bold=True, color='FFFFFF', size=11)
    fill_cat_h  = PatternFill('solid', fgColor='1E5F8A')
    font_cat_h  = Font(bold=True, color='DBEAFE', size=9)
    fill_aut_h  = PatternFill('solid', fgColor='065F46')
    font_aut_h  = Font(bold=True, color='D1FAE5', size=9)
    fill_sub    = PatternFill('solid', fgColor='EFF6FF')
    font_sub    = Font(bold=True, size=9, color='1E3A8A')
    fill_total  = PatternFill('solid', fgColor='0C4566')
    font_total  = Font(bold=True, color='FFFFFF', size=11)
    fill_alt    = PatternFill('solid', fgColor='F8F9FB')
    fill_alt_g  = PatternFill('solid', fgColor='F0FDF4')
    side        = Side(style='thin', color='D8DCE3')
    brd         = Border(left=side, right=side, top=side, bottom=side)
    c_ctr       = Alignment(horizontal='center', vertical='center')
    c_rgt       = Alignment(horizontal='right',  vertical='center')
    c_lft       = Alignment(vertical='center')
    NUM_COLS    = 8

    def merge_write(r, val, fill, font, align=None):
        ws.merge_cells(f'A{r}:H{r}')
        cell = ws.cell(row=r, column=1, value=val)
        cell.fill = fill; cell.font = font
        cell.alignment = align or Alignment(vertical='center')

    row = 1
    merge_write(row, 'REPORTE CONSOLIDADO DE VALUACIONES', fill_exp, Font(bold=True, color='FFFFFF', size=14),
                Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[row].height = 28
    row += 1

    parts = []
    if desde or hasta:
        parts.append(f"Período: {desde or '—'} a {hasta or '—'}")
    if exp_q:
        parts.append(f"Expediente: {exp_q}")
    if usuario:
        parts.append(f"Usuario: {usuario}")
    parts.append(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    merge_write(row, '  |  '.join(parts), PatternFill('solid', fgColor='F0F5F8'),
                Font(italic=True, size=9, color='475569'),
                Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[row].height = 16
    row += 2

    for g in grupos:
        merge_write(row,
                    f"EXPEDIENTE {g['expediente'] or '—'}  —  {g['caratula'] or '—'}  (Creado por: {g['creado_por'] or '—'})",
                    fill_exp, font_exp)
        ws.row_dimensions[row].height = 22
        row += 1

        if g['catastros']:
            for ci, h in enumerate(['VR #', 'Catastro', 'Tipo', 'Dirección', 'Total U$S', 'Propuesta U$S', 'Fecha Reg.', 'Reconside.'], 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.fill = fill_cat_h; cell.font = font_cat_h
                cell.alignment = c_ctr; cell.border = brd
            ws.row_dimensions[row].height = 18; row += 1

            for i, c in enumerate(g['catastros']):
                alt = fill_alt if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
                row_vals = [c['numero_vr'], c['catastro'], c['tipo_catastro'], c['direccion'],
                            c['total_usd'], c['propuesta'],
                            to_ar(c['fecha_registro']).strftime('%d/%m/%Y %H:%M') if c['fecha_registro'] else None,
                            'Sí' if c['es_reconsideracion'] else '']
                row_aligns = ['center','left','center','left','right','right','center','center']
                for ci, (val, al) in enumerate(zip(row_vals, row_aligns), 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.fill = alt; cell.border = brd; cell.font = Font(size=9)
                    cell.alignment = c_ctr if al == 'center' else (c_rgt if al == 'right' else c_lft)
                    if ci in (5, 6) and val is not None:
                        cell.number_format = '#,##0.00'
                row += 1

        if g['automotores']:
            for ci, h in enumerate(['Rec #', 'Vehículo', 'Año', 'Valor $', 'Creado por', 'Fecha Reg.', '', ''], 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.fill = fill_aut_h; cell.font = font_aut_h
                cell.alignment = c_ctr; cell.border = brd
            ws.row_dimensions[row].height = 18; row += 1

            for i, a in enumerate(g['automotores']):
                alt = fill_alt_g if i % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
                row_vals = [a['numero_recuento'], a['vehiculo'], a['anio'], a['valor'],
                            a['creado_por'],
                            to_ar(a['fecha_registro']).strftime('%d/%m/%Y %H:%M') if a['fecha_registro'] else None,
                            '', '']
                row_aligns = ['center','left','center','right','center','center','','']
                for ci, (val, al) in enumerate(zip(row_vals, row_aligns), 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.fill = alt; cell.border = brd; cell.font = Font(size=9)
                    cell.alignment = c_ctr if al == 'center' else (c_rgt if al == 'right' else c_lft)
                    if ci == 4 and val is not None:
                        cell.number_format = '#,##0'
                row += 1

        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=1, value='SUBTOTAL').fill = fill_sub
        ws.cell(row=row, column=1).font = font_sub
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        for ci, (val, fmt) in enumerate([(g['subtotal_usd'], '#,##0.00'), (g['subtotal_pesos'], '#,##0')], 5):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = fill_sub; cell.font = font_sub
            cell.alignment = c_rgt; cell.border = brd
            cell.number_format = fmt
        ws.row_dimensions[row].height = 18
        row += 2


    for ci, w in enumerate([8, 20, 14, 32, 16, 16, 18, 10], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fecha = datetime.now().strftime('%Y%m%d')
    return Response(buf,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="reporte_{fecha}.xlsx"'})


@app.route('/reporte/pdf')
@login_required
def reporte_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    desde   = request.args.get('desde',   '').strip()
    hasta   = request.args.get('hasta',   '').strip()
    exp_q   = request.args.get('exp_q',   '').strip()
    usuario = request.args.get('usuario', '').strip()

    conn = get_db()
    grupos, total_usd, total_pesos = _query_reporte(conn, desde, hasta, exp_q, usuario)
    conn.close()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            rightMargin=1.5*cm, leftMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles  = getSampleStyleSheet()
    brand   = colors.HexColor('#0C4566')
    brand2  = colors.HexColor('#1E5F8A')
    grn     = colors.HexColor('#065F46')
    gray_bg = colors.HexColor('#F1F5F9')
    grn_bg  = colors.HexColor('#F0FDF4')
    gray_ln = colors.HexColor('#CBD5E1')

    def _ar(v):
        return f'{float(v or 0):,.0f}'.replace(',', '.')

    def _ar_p(v):
        return f'{int(v or 0):,}'.replace(',', '.')

    s_title = ParagraphStyle('t', parent=styles['Heading1'], fontSize=13,
                              textColor=brand, alignment=TA_CENTER, spaceAfter=3)
    s_sub   = ParagraphStyle('s', parent=styles['Normal'], fontSize=8,
                              textColor=colors.HexColor('#64748B'), alignment=TA_CENTER, spaceAfter=10)
    s_exp   = ParagraphStyle('e', parent=styles['Normal'], fontSize=9.5, fontName='Helvetica-Bold',
                              textColor=colors.white, backColor=brand,
                              borderPadding=(4, 6, 4, 6), spaceBefore=10, spaceAfter=2)
    s_sec   = ParagraphStyle('sec', parent=styles['Normal'], fontSize=7.5,
                              textColor=colors.HexColor('#64748B'), spaceBefore=4, spaceAfter=2)
    s_stot  = ParagraphStyle('st', parent=styles['Normal'], fontSize=8.5,
                              textColor=colors.HexColor('#1E3A8A'), alignment=TA_RIGHT, spaceAfter=4)
    s_tot   = ParagraphStyle('tot', parent=styles['Heading2'], fontSize=11,
                              textColor=brand, alignment=TA_RIGHT, spaceBefore=6)

    def mk_table(header, rows, col_widths, hdr_color):
        data = [header] + rows
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, 0),  hdr_color),
            ('TEXTCOLOR',    (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, -1), 7.5),
            ('GRID',         (0, 0), (-1, -1), 0.4, gray_ln),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, gray_bg]),
            ('TOPPADDING',   (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 3),
            ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        return t

    story = []
    story.append(Paragraph('REPORTE CONSOLIDADO DE VALUACIONES', s_title))

    parts = []
    if desde or hasta:
        parts.append(f"Período: {desde or '—'} → {hasta or '—'}")
    if exp_q:
        parts.append(f"Expediente: {exp_q}")
    if usuario:
        parts.append(f"Usuario: {usuario}")
    parts.append(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    story.append(Paragraph('  |  '.join(parts), s_sub))

    if not grupos:
        story.append(Paragraph('No se encontraron registros para los filtros seleccionados.', styles['Normal']))
    else:
        for g in grupos:
            story.append(Paragraph(
                f"EXPEDIENTE {g['expediente'] or '—'}  —  {g['caratula'] or '—'}  "
                f"<font size='8'>(Creado por: {g['creado_por'] or '—'})</font>", s_exp))

            if g['catastros']:
                story.append(Paragraph('Inmuebles / Catastros', s_sec))
                cat_rows = []
                for c in g['catastros']:
                    rec = ' (R)' if c['es_reconsideracion'] else ''
                    cat_rows.append([
                        str(c['numero_vr'] or '—') + rec,
                        (c['catastro'] or '—'),
                        (c['tipo_catastro'] or '—'),
                        (c['direccion'] or '—')[:50],
                        f"U$S {_ar(c['total_usd'])}",
                        f"U$S {_ar(c['propuesta'])}",
                        to_ar(c['fecha_registro']).strftime('%d/%m/%y') if c['fecha_registro'] else '—',
                    ])
                story.append(mk_table(
                    ['VR #', 'Catastro', 'Tipo', 'Dirección', 'Total U$S', 'Propuesta U$S', 'Fecha Reg.'],
                    cat_rows,
                    [1.8*cm, 3.5*cm, 2.5*cm, 8.5*cm, 3*cm, 3.2*cm, 2.5*cm],
                    brand2
                ))

            if g['automotores']:
                story.append(Spacer(1, 0.2*cm))
                story.append(Paragraph('Automotores', s_sec))
                aut_rows = []
                for a in g['automotores']:
                    aut_rows.append([
                        str(a['numero_recuento'] or '—'),
                        (a['vehiculo'] or '—')[:45],
                        str(a['anio'] or '—'),
                        f"$ {_ar_p(a['valor'])}",
                        a['creado_por'] or '—',
                        to_ar(a['fecha_registro']).strftime('%d/%m/%y') if a['fecha_registro'] else '—',
                        '',
                    ])
                aut_t = mk_table(
                    ['Rec #', 'Vehículo', 'Año', 'Valor $', 'Creado por', 'Fecha Reg.', ''],
                    aut_rows,
                    [1.8*cm, 9*cm, 2*cm, 3.5*cm, 2.5*cm, 2.5*cm, 3.7*cm],
                    grn
                )
                aut_t.setStyle(TableStyle([
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, grn_bg]),
                ]))
                story.append(aut_t)

            sub_parts = []
            if g['subtotal_usd']:
                sub_parts.append(f"Inmuebles (propuesta): <b>U$S {_ar(g['subtotal_usd'])}</b>")
            if g['subtotal_pesos']:
                sub_parts.append(f"Automotores: <b>$ {_ar_p(g['subtotal_pesos'])}</b>")
            story.append(Paragraph('Subtotal: ' + ('  |  '.join(sub_parts) if sub_parts else '—'), s_stot))
            story.append(HRFlowable(width='100%', thickness=0.5, color=gray_ln, spaceAfter=2))



    doc.build(story)
    buf.seek(0)
    fecha = datetime.now().strftime('%Y%m%d')
    return Response(buf, mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename="reporte_{fecha}.pdf"'})


init_db()

if __name__ == '__main__':
    app.run(debug=True, port=5002)
